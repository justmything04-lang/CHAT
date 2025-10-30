import random
from datetime import datetime, timedelta
import calendar
import openpyxl

def working_days(start, end):
    cur = start
    days = []
    while cur <= end:
        if cur.weekday() != 6:  # Monday–Saturday
            days.append(cur)
        cur += timedelta(days=1)
    return days

def create_master_list(n, id_start):
    students = []
    for i in range(n):
        uid = str(id_start + i)
        name = f"Student_{uid[-4:]}"
        linked = random.choice(["Yes", "No"])
        invited = "Yes" if linked == "Yes" or random.random() < 0.8 else "No"
        phone = f"+91{random.randint(7000000000,9999999999)}" if invited == "Yes" else ""
        chatid = uid if linked == "Yes" else ""
        students.append([name, uid, phone, chatid, linked, invited])
    return students

def save_master_workbook():
    wb = openpyxl.Workbook()

    # MasterList
    ws = wb.active
    ws.title = "MasterList"
    ws.append(["Name","Reg ID","ParentPhone","ParentChatId","ParentLinked","ParentInvited"])
    for row in create_master_list(30, 2000000000):
        ws.append(row)

    # OnlineMasterList
    ws2 = wb.create_sheet("OnlineMasterList")
    ws2.append(["Name","Reg ID","ParentPhone","ParentChatId","ParentLinked","ParentInvited"])
    for row in create_master_list(18, 3000000000):
        ws2.append(row)

    # Settings
    ws3 = wb.create_sheet("Settings")
    ws3.append(["DailyEasterEgg","StartTime","EndTime","ClassStartDate"])
    start_date = (datetime.today().replace(day=1) - timedelta(days=90)).date()
    ws3.append(["Have a nice day!", "09:00", "13:00", str(start_date)])

    wb.save("main.xlsx")

def save_absentee_workbook(fname, students, suffix):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    today = datetime.today().date()
    start = (today.replace(day=1) - timedelta(days=90))
    end = today
    days = working_days(start, end)

    for d in days:
        ws = wb.create_sheet(f"{d}-{suffix}")
        ws.append(["Name","Reg ID","Date"])
        absents = random.sample(students, k=random.randint(2, max(3, len(students)//5)))
        for s in absents:
            ws.append([s[0], s[1], str(d)])

    wb.save(fname)

if __name__ == "__main__":
    save_master_workbook()

    # Load students
    offline = create_master_list(30, 2000000000)
    online  = create_master_list(18, 3000000000)

    save_absentee_workbook("offline_absentees.xlsx", offline, "offline")
    save_absentee_workbook("online_absentees.xlsx", online, "online")

    print("✅ Dummy Excel sheets created: main.xlsx, offline_absentees.xlsx, online_absentees.xlsx")
